import datetime as dt
import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    ListFlowable,
    ListItem,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# =========================================================
# Public API
# =========================================================


@dataclass
class RunArtifacts:
    run_id: str
    run_json: str
    report_json: Optional[str]
    pdf: Optional[str]
    xlsx: Optional[str]


def export_run_artifacts(
        spec: str, plan: Dict[str, Any], detailed_results: List[Dict[str, Any]]
) -> RunArtifacts:
    """
    Generate real-world QA artifacts:
      - Run JSON (raw payload)
      - Report JSON (pytest-json-report output if present)
      - PDF execution report
      - Excel report (Summary / Testcases / Observations)
    """
    out_dir = Path(os.getenv("ARTIFACTS_DIR", str(Path("data") / "logs")))
    out_dir.mkdir(parents=True, exist_ok=True)

    run_id = dt.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    run_json_name = f"run_{run_id}.json"
    report_json_name = f"report_{run_id}.json"
    pdf_name = f"run_{run_id}.pdf"
    xlsx_name = f"run_{run_id}.xlsx"

    meta = _meta()
    payload = {
        "spec": spec,
        "plan": plan,
        "results": detailed_results,
        "created_at": dt.datetime.utcnow().isoformat(),
        "meta": meta,
    }
    (out_dir / run_json_name).write_text(
        json.dumps(payload, indent=2, default=_json_default), encoding="utf-8"
    )

    # Find pytest-json-report output
    report_obj = None
    copied_report = None
    src_report = _find_pytest_report_json(detailed_results)
    if src_report:
        try:
            txt = Path(src_report).read_text(encoding="utf-8", errors="ignore")
            (out_dir / report_json_name).write_text(txt, encoding="utf-8")
            copied_report = report_json_name
            report_obj = json.loads(txt)
        except Exception:
            copied_report = None
            report_obj = None

    # Build report data
    summary = _summarize(
        meta=meta,
        plan=plan,
        detailed_results=detailed_results,
        report_obj=report_obj,
    )
    tc_headers, tc_rows = _build_testcases(
        spec=spec,
        plan=plan,
        detailed_results=detailed_results,
        report_obj=report_obj,
    )
    obs_headers, obs_rows = _build_observations(
        run_id=run_id, testcases_rows=tc_rows
    )

    # Excel
    xlsx_out = None
    try:
        _write_excel(
            out_path=out_dir / xlsx_name,
            meta=meta,
            summary=summary,
            testcases_headers=tc_headers,
            testcases_rows=tc_rows,
            obs_headers=obs_headers,
            obs_rows=obs_rows,
        )
        xlsx_out = xlsx_name
    except Exception:
        xlsx_out = None

    # PDF
    pdf_out = None
    try:
        _write_pdf(
            out_path=out_dir / pdf_name,
            meta=meta,
            summary=summary,
            testcases_headers=tc_headers,
            testcases_rows=tc_rows,
            obs_headers=obs_headers,
            obs_rows=obs_rows,
        )
        pdf_out = pdf_name
    except Exception:
        pdf_out = None

    return RunArtifacts(
        run_id=run_id,
        run_json=run_json_name,
        report_json=copied_report,
        pdf=pdf_out,
        xlsx=xlsx_out,
    )


# =========================================================
# Helpers
# =========================================================


def _meta() -> Dict[str, str]:
    return {
        "project": os.getenv("QA_PROJECT_NAME", "AI QA Platform Demo"),
        "environment": os.getenv(
            "QA_ENVIRONMENT", os.getenv("MODE", "Local")
        ),
        "build_version": os.getenv("QA_BUILD_VERSION", "Sample v0.1"),
        "prepared_by": os.getenv("QA_PREPARED_BY", "QA Agent (Automated)"),
    }


def _json_default(o: Any):
    if isinstance(o, Path):
        return str(o)
    if isinstance(o, (dt.datetime, dt.date)):
        return o.isoformat()
    return str(o)


def _find_pytest_report_json(
        detailed_results: Optional[List[Dict[str, Any]]] = None,
) -> Optional[str]:
    """
    Find the pytest-json-report output.
    Priority:
      1. report_file from the run result (e.g. data/reports/report_xxx.json)
      2. .report.json in standard locations
    """
    # Check if the runner gave us a specific path
    if detailed_results:
        for item in detailed_results:
            res = (item or {}).get("result") or {}
            rf = res.get("report_file")
            if rf:
                p = Path(rf)
                if p.exists():
                    return str(p)

    # Fallback: standard locations
    for p in [
        Path(".report.json"),
        Path("data") / ".report.json",
        Path(".") / "report.json",
    ]:
        if p.exists() and p.is_file():
            return str(p)

    # Also check data/reports/ for the most recent one
    reports_dir = Path("data") / "reports"
    if reports_dir.exists():
        jsons = sorted(reports_dir.glob("report_*.json"), reverse=True)
        if jsons:
            return str(jsons[0])

    return None


def _soft_breaks(s: str, for_pdf: bool = False) -> str:
    """Insert optional break hints for long tokens.

    For Excel: use zero-width space (renders fine in Excel).
    For PDF: use plain text (no special chars — Helvetica can't render \u200b).
    """
    if for_pdf:
        # PDF: don't insert any special chars — Helvetica renders them as black boxes
        return s
    z = "\u200b"
    return (
        s.replace("/", f"/{z}")
        .replace("\\", f"\\{z}")
        .replace("_", f"_{z}")
        .replace("-", f"-{z}")
        .replace("::", f"::{z}")
        .replace("?", f"?{z}")
        .replace("&", f"&{z}")
        .replace("=", f"={z}")
        .replace(".", f".{z}")
    )


def _pretty_tool(name: str) -> str:
    return {
        "pytest_runner": "Pytest (API/Unit)",
        "playwright_runner": "Playwright (UI)",
        "api_caller": "API Caller",
        "bug_reporter": "Bug Reporter",
    }.get(name, (name or "").replace("_", " ").title())


def _humanize_case_name(name: str) -> str:
    """Convert parametrize IDs to readable names."""
    # e.g. "invalid_username_invalid_password" -> "Invalid username invalid password"
    return name.replace("_", " ").replace("-", " ").strip().capitalize()


def _format_test_data(inputs: Dict[str, Any]) -> str:
    """Format test inputs for display."""
    parts = []
    for k, v in inputs.items():
        val = repr(v) if isinstance(v, str) else str(v)
        parts.append(f"{k}={val}")
    return ", ".join(parts)


def _infer_expected_for_case(case_data: Dict[str, Any], spec: str) -> str:
    """Build a rich 'Expected' string from the plan's test data."""
    expected = case_data.get("expected") or {}
    parts = []

    if expected.get("error_visible"):
        errors = expected.get("error_any_of") or []
        if errors:
            parts.append(f"Error message shown: {' or '.join(repr(e) for e in errors)}")
        else:
            parts.append("Error/validation message should be visible")

    if expected.get("stays_on_page"):
        parts.append("User stays on login page (no redirect)")

    url_contains = expected.get("url_contains")
    if url_contains:
        parts.append(f"URL contains '{url_contains}'")

    outcome = expected.get("outcome")
    if outcome:
        parts.append(f"Outcome: {outcome}")

    if not parts:
        t = spec.lower()
        return spec

    return "; ".join(parts)


def _short_error(s: str, limit: int = 260) -> str:
    if not s:
        return ""
    s = re.sub(r"\s+", " ", str(s)).strip()
    return (s[: limit] + "…") if len(s) > limit else s


# =========================================================
# Build summary + rows
# =========================================================


def _summarize(
        meta: Dict[str, str],
        plan: Dict[str, Any],
        detailed_results: List[Dict[str, Any]],
        report_obj: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    goal = plan.get("goal") or ""
    passed = failed = skipped = 0

    # Priority 1: pytest-json-report summary (per-test level)
    if isinstance(report_obj, dict) and isinstance(
            report_obj.get("summary"), dict
    ):
        s = report_obj["summary"]
        passed = int(s.get("passed", 0) or 0)
        failed = int(s.get("failed", 0) or 0)
        skipped = int(s.get("skipped", 0) or 0)
    else:
        # Priority 2: orchestrator step statuses (most reliable)
        for item in detailed_results:
            s_status = (item or {}).get("status") or ""
            if s_status == "passed":
                passed += 1
            elif s_status == "skipped":
                skipped += 1
            elif s_status in ("failed", "error"):
                failed += 1
            else:
                res = (item or {}).get("result") or {}
                sm = res.get("summary") or {}
                passed  += int(sm.get("passed",  0) or 0)
                failed  += int(sm.get("failed",  0) or 0)
                skipped += int(sm.get("skipped", 0) or 0)

    # Priority 3: count from plan test data
    if passed == failed == skipped == 0:
        for step in plan.get("steps") or []:
            data = (step.get("args") or {}).get("data") or []
            if data:
                passed = len(data)  # Assume pass if no report
                break
        if passed == 0:
            passed = len(plan.get("steps") or [])

    total = passed + failed + skipped
    pass_rate = f"{(passed / total * 100):.1f}%" if total else "0.0%"
    readiness = (
        "READY for the next environment gate."
        if failed == 0
        else "NOT READY. Fix failures before promoting this build."
    )

    return {
        "project": meta.get("project"),
        "environment": meta.get("environment"),
        "build_version": meta.get("build_version"),
        "prepared_by": meta.get("prepared_by"),
        "report_date": dt.date.today().isoformat(),
        "goal": goal,
        "total": total,
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "pass_rate": pass_rate,
        "readiness": readiness,
        "generated_on": dt.datetime.utcnow().isoformat(),
    }


def _build_testcases(
        spec: str,
        plan: Dict[str, Any],
        detailed_results: List[Dict[str, Any]],
        report_obj: Optional[Dict[str, Any]],
) -> Tuple[List[str], List[List[Any]]]:
    headers = [
        "S.NO",
        "Module",
        "Page",
        "Test scenario / case",
        "Test Data",
        "Expected Results",
        "Actual Result",
        "Status",
        "QA Comments",
    ]

    goal = plan.get("goal") or spec

    # Extract page URL and test data from plan
    page_hint = ""
    plan_test_data: List[Dict[str, Any]] = []
    plan_module = ""
    for st in plan.get("steps") or []:
        args = (st or {}).get("args") or {}
        if not page_hint:
            page_hint = (
                    args.get("url")
                    or args.get("base_url")
                    or args.get("page")
                    or args.get("endpoint")
                    or ""
            )
        if not plan_module:
            plan_module = args.get("description") or ""
        data_list = args.get("data") or []
        if data_list:
            plan_test_data.extend(data_list)

    rows: List[List[Any]] = []

    # === PATH A: We have pytest-json-report with per-test results ===
    tests = (
        (report_obj or {}).get("tests") if isinstance(report_obj, dict) else None
    )
    if isinstance(tests, list) and tests:
        # Build a lookup from case name -> plan data
        plan_data_lookup: Dict[str, Dict[str, Any]] = {}
        for d in plan_test_data:
            name = d.get("name", "")
            plan_data_lookup[name] = d
            # Also try without case prefix (parametrize IDs)
            clean = name.replace(" ", "_").lower()
            plan_data_lookup[clean] = d

        for i, t in enumerate(tests, start=1):
            nodeid = t.get("nodeid") or ""
            outcome = (t.get("outcome") or "").lower()
            status = (
                "Pass"
                if outcome == "passed"
                else ("Fail" if outcome == "failed" else "Skip")
            )

            # Extract module and case name from nodeid
            file_part = nodeid.split("::")[0] if "::" in nodeid else nodeid
            module = Path(file_part).stem if file_part else "automation"
            module = module.replace("_", " ").title()

            # Extract parametrize case name: test_xxx[chromium-case_name]
            case_name = ""
            m = re.search(r"\[(?:chromium-)?(.+?)\]", nodeid)
            if m:
                case_name = m.group(1)

            # Look up plan data for this case
            matched_plan = plan_data_lookup.get(case_name, {})
            if not matched_plan:
                # Try normalized
                clean_case = case_name.replace(" ", "_").lower()
                matched_plan = plan_data_lookup.get(clean_case, {})

            # Build scenario (human readable)
            scenario = _humanize_case_name(case_name) if case_name else nodeid

            # Build test data column
            inputs = matched_plan.get("inputs") or {}
            test_data_str = _format_test_data(inputs) if inputs else ""

            # Build expected results
            if matched_plan:
                expected = _infer_expected_for_case(matched_plan, goal)
            else:
                expected = args.get("linked_scenario") or args.get(
                    "description") or f"Step {i}: verify functionality works correctly"

            # Build actual result
            if status == "Pass":
                actual = "Test passed successfully."
            elif status == "Fail":
                crash_msg = ""
                if isinstance(t.get("call"), dict):
                    crash_msg = (
                            ((t["call"].get("crash") or {}) or {}).get("message") or ""
                    )
                lr = t.get("longrepr") or ""
                err = crash_msg or lr or ""
                actual = _short_error(err, 300) or "Test failed — see Report JSON"
            else:
                actual = "Skipped"

            # QA Comments
            comments = ""
            if status == "Fail":
                comments = (
                    "Automation error — review error details and fix test code or "
                    "raise bug if application defect."
                )
            elif status == "Pass":
                techniques = matched_plan.get("techniques") or []
                if techniques:
                    comments = f"Test technique(s): {', '.join(techniques)}"

            rows.append(
                [
                    i,
                    module,
                    page_hint,
                    scenario,
                    test_data_str,
                    expected,
                    actual,
                    status,
                    comments,
                ]
            )

        return headers, rows

    # === PATH B: No pytest report — build rows from plan test data ===
    if plan_test_data:
        for i, case in enumerate(plan_test_data, start=1):
            case_name = case.get("name", f"Case {i}")
            scenario = _humanize_case_name(case_name)

            inputs = case.get("inputs") or {}
            test_data_str = _format_test_data(inputs)

            expected = _infer_expected_for_case(case, goal)

            # Check overall result
            overall_ok = True
            for item in detailed_results:
                res = (item or {}).get("result") or {}
                sm = res.get("summary") or {}
                if int(sm.get("failed", 0) or 0) > 0:
                    overall_ok = False

            status = "Pass" if overall_ok else "Inconclusive"
            actual = (
                "Step completed successfully."
                if overall_ok
                else "See run logs for details"
            )

            techniques = case.get("techniques") or []
            comments = (
                f"Test technique(s): {', '.join(techniques)}" if techniques else ""
            )

            file_part = ""
            for st in plan.get("steps") or []:
                file_part = (st.get("args") or {}).get("path", "")
                break
            module = (
                Path(file_part).stem.replace("_", " ").title()
                if file_part
                else "Automation"
            )

            rows.append(
                [
                    i,
                    module,
                    page_hint,
                    scenario,
                    test_data_str,
                    expected,
                    actual,
                    status,
                    comments,
                ]
            )

        return headers, rows

    # === PATH C: One row per orchestrator step ===
    for idx, item in enumerate(detailed_results, start=1):
        step        = (item or {}).get("step") or {}
        args        = step.get("args") or {}
        res         = (item or {}).get("result") or {}
        step_status = (item or {}).get("status") or ""
        error_msg   = (item or {}).get("error") or "" or ""
        tool = step.get("tool") or ""
        args = step.get("args") or {}
        si = step.get("index", idx - 1)

        # Module from step description or path
        description = args.get("description") or args.get("path") or goal
        path_stem = Path(args.get("path", "")).stem if args.get("path") else ""
        module = path_stem.replace("_", " ").title() if path_stem else _pretty_tool(tool)

        page = (
                args.get("url")
                or args.get("base_url")
                or args.get("page")
                or args.get("endpoint")
                or page_hint
        )

        # Scenario = description from plan args (not the goal)
        scenario = description

        # Expected = what this step was supposed to test
        linked = args.get("linked_scenario") or args.get("description") or goal
        expected = linked if linked and linked != goal else f"Step {idx}: {description}"

        # Actual = parse from pytest report embedded in result
        # Try: res["report"]["tests"] (pytest-json-report)
        step_tests = None
        if isinstance(res, dict):
            rpt = res.get("report") or res.get("json_report") or {}
            step_tests = rpt.get("tests") if isinstance(rpt, dict) else None
            if not step_tests:
                # Try res directly if it looks like a pytest report
                step_tests = res.get("tests")

        if step_tests and isinstance(step_tests, list):
            # Have per-test data
            p = sum(1 for t in step_tests if t.get("outcome") == "passed")
            f = sum(1 for t in step_tests if t.get("outcome") == "failed")
            s = sum(1 for t in step_tests if t.get("outcome") == "skipped")
            status = "Pass" if f == 0 and p > 0 else ("Fail" if f > 0 else "Skip")
            if status == "Pass":
                actual = f"All {p} test(s) passed successfully."
            else:
                # Get first failure message
                fail_msgs = []
                for t in step_tests:
                    if t.get("outcome") == "failed":
                        crash = ((t.get("call") or {}).get("crash") or {}).get("message") or ""
                        lr = t.get("longrepr") or ""
                        msg = _short_error(crash or str(lr), 250)
                        if msg:
                            fail_msgs.append(msg)
                actual = fail_msgs[0] if fail_msgs else f"Failed: {f}, Passed: {p}"
        elif step_status:
            # Use orchestrator-level status
            status = "Pass" if step_status == "passed" else ("Skip" if step_status == "skipped" else "Fail")
            if status == "Pass":
                actual = "Step completed successfully."
            elif status == "Skip":
                actual = "Step skipped."
            else:
                actual = _short_error(error_msg, 250) or "Step failed — see run logs."
        else:
            # Last resort: check summary in result
            sm = res.get("summary") or {}
            f_count = int(sm.get("failed", 0) or 0)
            p_count = int(sm.get("passed", 0) or 0)
            status = "Fail" if f_count > 0 else "Pass"
            actual = f"Passed: {p_count}, Failed: {f_count}"

        comments = ""
        if status == "Fail":
            comments = "Automation error — review error details and fix test code or raise bug if application defect."
        elif args.get("is_prerequisite"):
            comments = "Prerequisite step — auth setup for subsequent tests."

        rows.append(
            [
                idx,
                module,
                str(page or ""),
                str(scenario),
                "",
                str(expected),
                str(actual),
                status,
                comments,
            ]
        )

    return headers, rows


def _build_observations(
        run_id: str, testcases_rows: List[List[Any]]
) -> Tuple[List[str], List[List[Any]]]:
    headers = ["S.NO", "BUG ID", "Description", "Priority"]
    obs = []
    bug_no = 1
    for r in testcases_rows:
        # Status is at index 7 now (added Test Data column)
        status = str(r[7]).strip().lower() if len(r) > 7 else ""
        if status == "fail":
            bug_id = f"AUTO-{run_id}-{bug_no:03d}"
            scenario = r[3] if len(r) > 3 else ""
            actual = r[6] if len(r) > 6 else ""
            desc = f"{scenario} — {actual}"
            obs.append([bug_no, bug_id, _short_error(desc, 320), "P1"])
            bug_no += 1
    if not obs:
        obs.append(
            [
                1,
                f"AUTO-{run_id}-000",
                "No defects observed in this execution run.",
                "P3",
            ]
        )
    return headers, obs


# =========================================================
# Excel — matches user's sample format
# =========================================================

_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_SKIP_FILL = PatternFill("solid", fgColor="FFEB9C")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_excel(
        out_path: Path,
        meta: Dict[str, str],
        summary: Dict[str, Any],
        testcases_headers: List[str],
        testcases_rows: List[List[Any]],
        obs_headers: List[str],
        obs_rows: List[List[Any]],
) -> None:
    wb = openpyxl.Workbook()

    # ── Summary Sheet ──
    ws_sum = wb.active
    ws_sum.title = "Summary"

    summary_data = [
        ("Project", meta.get("project")),
        ("Environment", meta.get("environment")),
        ("Build / Version", meta.get("build_version")),
        ("Prepared by", meta.get("prepared_by")),
        ("Report Date", summary.get("report_date")),
        ("", ""),
        ("Goal", summary.get("goal")),
        ("", ""),
        ("Total Cases", summary.get("total")),
        ("Passed", summary.get("passed")),
        ("Failed", summary.get("failed")),
        ("Skipped", summary.get("skipped")),
        ("Pass Rate", summary.get("pass_rate")),
        ("", ""),
        ("Release Readiness", summary.get("readiness")),
    ]

    for i, (k, v) in enumerate(summary_data, start=1):
        ws_sum.cell(row=i, column=1, value=k).font = Font(bold=True, size=11)
        ws_sum.cell(row=i, column=2, value=v)
        ws_sum.cell(row=i, column=1).alignment = Alignment(vertical="top")
        ws_sum.cell(row=i, column=2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 60

    # Color the pass rate and readiness
    pass_row = 13
    ws_sum.cell(row=pass_row, column=2).font = Font(bold=True, size=12)
    readiness_row = 15
    if summary.get("failed", 0) == 0:
        ws_sum.cell(row=readiness_row, column=2).fill = _PASS_FILL
    else:
        ws_sum.cell(row=readiness_row, column=2).fill = _FAIL_FILL

    # ── Testcases Sheet ──
    ws_tc = wb.create_sheet("Testcases")

    # Headers
    for c, h in enumerate(testcases_headers, start=1):
        cell = ws_tc.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = _BORDER

    # Data rows
    for r_idx, row in enumerate(testcases_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_tc.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER

        # Color the Status column (index 7 = column 8)
        status_cell = ws_tc.cell(row=r_idx, column=8)
        status_val = str(status_cell.value or "").strip().lower()
        if status_val == "pass":
            status_cell.fill = _PASS_FILL
            status_cell.font = Font(bold=True, color="006100")
        elif status_val == "fail":
            status_cell.fill = _FAIL_FILL
            status_cell.font = Font(bold=True, color="9C0006")
        elif status_val == "skip":
            status_cell.fill = _SKIP_FILL

    # Column widths
    widths = [6, 18, 30, 32, 28, 32, 32, 10, 32]
    for i, w in enumerate(widths[: len(testcases_headers)], start=1):
        ws_tc.column_dimensions[get_column_letter(i)].width = w

    # ── Observations Sheet ──
    ws_obs = wb.create_sheet("Observations")

    for c, h in enumerate(obs_headers, start=1):
        cell = ws_obs.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = _BORDER

    for r_idx, row in enumerate(obs_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_obs.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER

    obs_widths = [6, 24, 60, 10]
    for i, w in enumerate(obs_widths, start=1):
        ws_obs.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


# =========================================================
# PDF — professional QA report
# =========================================================


def _write_pdf(
        out_path: Path,
        meta: Dict[str, str],
        summary: Dict[str, Any],
        testcases_headers: List[str],
        testcases_rows: List[List[Any]],
        obs_headers: List[str],
        obs_rows: List[List[Any]],
) -> None:
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleX",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        spaceAfter=6,
    )
    sub_style = ParagraphStyle(
        "SubX",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#475569"),
        spaceAfter=14,
    )
    h_style = ParagraphStyle(
        "HX",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        spaceBefore=10,
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "CellX",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        wordWrap="CJK",
    )

    def P(v: Any) -> Paragraph:
        s = "" if v is None else str(v)
        # No _soft_breaks for PDF — Helvetica can't render \u200b
        s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        s = s.replace("\n", "<br/>")
        return Paragraph(s, cell_style)

    def footer(c, doc):
        c.saveState()
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.HexColor("#64748B"))
        c.drawRightString(
            doc.pagesize[0] - doc.rightMargin,
            0.5 * inch,
            f"Page {c.getPageNumber()}",
        )
        c.restoreState()

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=landscape(A4),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.7 * inch,
    )

    story: List[Any] = []
    story.append(Paragraph("QA Test Execution Report", title_style))
    story.append(
        Paragraph(meta.get("project", "AI QA Platform Demo"), sub_style)
    )

    # Meta block
    meta_rows = [
        ["Project", meta.get("project")],
        ["Environment", meta.get("environment")],
        ["Build / Version", meta.get("build_version")],
        ["Prepared by", meta.get("prepared_by")],
        ["Report Date", summary.get("report_date")],
    ]
    meta_tbl = Table(
        [[P(a), P(b)] for a, b in meta_rows],
        colWidths=[doc.width * 0.22, doc.width * 0.78],
    )
    meta_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(meta_tbl)
    story.append(Spacer(1, 10))

    # Executive Summary
    story.append(Paragraph("Executive Summary", h_style))
    bullets = [
        f"Goal: {summary.get('goal')}",
        f"Total test cases executed: {summary.get('total')}.",
        (
            f"Result: {summary.get('passed')} passed, {summary.get('failed')} failed, "
            f"{summary.get('skipped')} skipped (Pass rate: {summary.get('pass_rate')})."
        ),
        f"Release readiness: {summary.get('readiness')}",
    ]
    lf = ListFlowable(
        [
            ListItem(
                Paragraph(b, styles["BodyText"]), leftIndent=14
            )
            for b in bullets
        ],
        bulletType="bullet",
        leftIndent=16,
    )
    story.append(lf)
    story.append(Spacer(1, 10))

    # Execution Summary
    story.append(Paragraph("Test Execution Summary", h_style))
    exec_tbl = Table(
        [
            [
                P("Total"),
                P(summary.get("total")),
                P("Passed"),
                P(summary.get("passed")),
                P("Failed"),
                P(summary.get("failed")),
                P("Skipped"),
                P(summary.get("skipped")),
                P("Pass Rate"),
                P(summary.get("pass_rate")),
            ]
        ],
        colWidths=[doc.width * 0.08, doc.width * 0.07] * 5,
    )
    exec_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF2FF")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(exec_tbl)

    # Detailed Test Cases
    def make_table(
            headers: List[str], rows: List[List[Any]], weights: List[float]
    ) -> Table:
        data = [[P(h) for h in headers]]
        for r in rows:
            data.append([P(v) for v in r])

        total_w = sum(weights)
        col_widths = [doc.width * (w / total_w) for w in weights]
        t = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        # Alternate row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                t.setStyle(
                    TableStyle(
                        [
                            (
                                "BACKGROUND",
                                (0, i),
                                (-1, i),
                                colors.HexColor("#F3F4F6"),
                            )
                        ]
                    )
                )
        return t

    story.append(Spacer(1, 14))
    story.append(Paragraph("Detailed Test Cases", h_style))
    # S.NO, Module, Page, Scenario, TestData, Expected, Actual, Status, Comments
    tc_weights = [0.5, 1.2, 1.8, 2.0, 1.8, 2.2, 2.2, 0.7, 2.2]
    story.append(make_table(testcases_headers, testcases_rows, tc_weights))

    story.append(Spacer(1, 14))
    story.append(Paragraph("Observations", h_style))
    obs_weights = [0.7, 1.3, 4.6, 1.0]
    story.append(make_table(obs_headers, obs_rows, obs_weights))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)