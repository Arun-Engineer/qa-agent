"""
Phase 6 · Workbook Writer
openpyxl Excel report generation with conditional formatting,
charts, and OneDrive sync-back support.
"""

import logging
import os
from typing import Any, Dict, List, Optional
from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── style constants ────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
PASS_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
WARN_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
SKIP_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

STATUS_FILLS = {"passed": PASS_FILL, "failed": FAIL_FILL, "skipped": SKIP_FILL, "warning": WARN_FILL}


class WorkbookWriter:
    """
    Generates rich Excel test reports with multiple sheets,
    conditional formatting, charts, and optional OneDrive upload.
    """

    def __init__(self, output_dir: str = "/tmp/reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate(
        self,
        run_data: Dict[str, Any],
        gate_decision: Optional[Dict] = None,
        filename: Optional[str] = None,
    ) -> str:
        """Generate Excel workbook. Returns file path."""
        wb = openpyxl.Workbook()

        # Sheet 1: Executive Summary
        self._write_summary_sheet(wb.active, run_data, gate_decision)

        # Sheet 2: Test Results
        ws_results = wb.create_sheet("Test Results")
        self._write_results_sheet(ws_results, run_data.get("results", []))

        # Sheet 3: Bugs
        ws_bugs = wb.create_sheet("Bugs")
        self._write_bugs_sheet(ws_bugs, run_data.get("bugs", []))

        # Sheet 4: Coverage
        ws_cov = wb.create_sheet("Coverage")
        self._write_coverage_sheet(ws_cov, run_data)

        # Sheet 5: Trends
        ws_trends = wb.create_sheet("Trends")
        self._write_trends_sheet(ws_trends, run_data.get("history", []))

        # Sheet 6: Gate Rules (if present)
        if gate_decision and gate_decision.get("rules"):
            ws_gate = wb.create_sheet("Release Gate")
            self._write_gate_sheet(ws_gate, gate_decision)

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        fname = filename or f"qa_report_{timestamp}.xlsx"
        path = os.path.join(self.output_dir, fname)
        wb.save(path)
        logger.info("Excel report: %s", path)
        return path

    # ── Sheet: Summary ─────────────────────────────────────────

    def _write_summary_sheet(self, ws, run_data: Dict, gate: Optional[Dict]):
        ws.title = "Summary"
        ws.sheet_properties.tabColor = "2E75B6"

        results = run_data.get("results", [])
        total = len(results)
        passed = sum(1 for r in results if r.get("status") == "passed")
        failed = sum(1 for r in results if r.get("status") == "failed")
        skipped = total - passed - failed

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "QA Test Execution Report"
        ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="2E75B6")

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Run: {run_data.get('run_id', 'N/A')} | Environment: {run_data.get('environment', 'N/A')} | {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A2"].font = Font(size=10, color="888888")

        # KPIs
        kpis = [
            ("Total Tests", total), ("Passed", passed), ("Failed", failed),
            ("Skipped", skipped), ("Pass Rate", f"{round(passed/total*100,1) if total else 0}%"),
            ("Bugs Filed", len(run_data.get("bugs", []))),
        ]
        for i, (label, value) in enumerate(kpis):
            col = i + 1
            cell_label = ws.cell(row=4, column=col, value=label)
            cell_label.font = Font(bold=True, size=10)
            cell_label.alignment = CENTER
            cell_val = ws.cell(row=5, column=col, value=value)
            cell_val.font = Font(bold=True, size=16)
            cell_val.alignment = CENTER
            if label == "Passed":
                cell_val.font = Font(bold=True, size=16, color="27AE60")
            elif label == "Failed":
                cell_val.font = Font(bold=True, size=16, color="E74C3C")

        # Gate verdict
        if gate:
            ws.cell(row=7, column=1, value="Release Gate:").font = Font(bold=True, size=11)
            v_cell = ws.cell(row=7, column=2, value=gate.get("verdict", "N/A"))
            v_cell.font = Font(bold=True, size=14,
                               color="27AE60" if gate.get("verdict") == "PASS" else
                               "E74C3C" if gate.get("verdict") == "FAIL" else "F39C12")
            ws.cell(row=7, column=3, value=f"Score: {gate.get('score', 0)}%").font = Font(size=11)

        # Pie chart
        if total > 0:
            chart_data_start = 10
            ws.cell(row=chart_data_start, column=1, value="Status")
            ws.cell(row=chart_data_start, column=2, value="Count")
            for i, (label, val) in enumerate([("Passed", passed), ("Failed", failed), ("Skipped", skipped)]):
                ws.cell(row=chart_data_start + 1 + i, column=1, value=label)
                ws.cell(row=chart_data_start + 1 + i, column=2, value=val)

            pie = PieChart()
            pie.title = "Test Results Distribution"
            pie.width = 16
            pie.height = 10
            labels = Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_start + 3)
            data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_start + 3)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            ws.add_chart(pie, "A15")

        # Column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet: Results ─────────────────────────────────────────

    def _write_results_sheet(self, ws, results: List[Dict]):
        ws.sheet_properties.tabColor = "27AE60"
        headers = ["#", "Test Name", "Module", "Category", "Status", "Duration (ms)", "Error", "Flaky"]
        self._write_header_row(ws, headers)

        for i, r in enumerate(results, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).alignment = CENTER
            ws.cell(row=row, column=2, value=r.get("name", "")[:80]).alignment = LEFT_WRAP
            ws.cell(row=row, column=3, value=r.get("module", "")).alignment = CENTER
            ws.cell(row=row, column=4, value=r.get("category", "")).alignment = CENTER
            status = r.get("status", "unknown")
            s_cell = ws.cell(row=row, column=5, value=status.upper())
            s_cell.alignment = CENTER
            s_cell.fill = STATUS_FILLS.get(status, PatternFill())
            s_cell.font = Font(bold=True)
            ws.cell(row=row, column=6, value=r.get("duration_ms", 0)).alignment = CENTER
            ws.cell(row=row, column=7, value=r.get("error", "")[:120]).alignment = LEFT_WRAP
            ws.cell(row=row, column=8, value="Yes" if r.get("flaky_count", 0) > 0 else "").alignment = CENTER

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = BORDER

        # Auto-fit columns
        widths = [5, 45, 15, 15, 10, 12, 50, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Auto-filter
        ws.auto_filter.ref = f"A1:H{len(results)+1}"

    # ── Sheet: Bugs ────────────────────────────────────────────

    def _write_bugs_sheet(self, ws, bugs: List[Dict]):
        ws.sheet_properties.tabColor = "E74C3C"
        headers = ["ID", "Title", "Severity", "Status", "Assigned To", "Module", "Created"]
        self._write_header_row(ws, headers)

        sev_fills = {
            "critical": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
            "high": PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid"),
            "medium": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
            "low": PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
        }
        sev_fonts = {k: Font(bold=True, color="FFFFFF") for k in sev_fills}

        for i, b in enumerate(bugs, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=b.get("id", "")).alignment = CENTER
            ws.cell(row=row, column=2, value=b.get("title", "")[:80]).alignment = LEFT_WRAP
            sev = b.get("severity", "medium").lower()
            s_cell = ws.cell(row=row, column=3, value=sev.upper())
            s_cell.fill = sev_fills.get(sev, PatternFill())
            s_cell.font = sev_fonts.get(sev, Font(bold=True))
            s_cell.alignment = CENTER
            ws.cell(row=row, column=4, value=b.get("status", "New")).alignment = CENTER
            ws.cell(row=row, column=5, value=b.get("assigned_to", "")).alignment = CENTER
            ws.cell(row=row, column=6, value=b.get("module", "")).alignment = CENTER
            ws.cell(row=row, column=7, value=b.get("created", "")).alignment = CENTER

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = BORDER

        widths = [12, 45, 12, 12, 20, 15, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = f"A1:G{len(bugs)+1}"

    # ── Sheet: Coverage ────────────────────────────────────────

    def _write_coverage_sheet(self, ws, run_data: Dict):
        ws.sheet_properties.tabColor = "2ECC71"
        headers = ["Module", "Total Tests", "Passed", "Failed", "Coverage %"]
        self._write_header_row(ws, headers)

        results = run_data.get("results", [])
        modules: Dict[str, Dict] = {}
        for r in results:
            mod = r.get("module", "Unknown")
            if mod not in modules:
                modules[mod] = {"total": 0, "passed": 0, "failed": 0}
            modules[mod]["total"] += 1
            if r.get("status") == "passed":
                modules[mod]["passed"] += 1
            elif r.get("status") == "failed":
                modules[mod]["failed"] += 1

        for i, (mod, d) in enumerate(sorted(modules.items()), 1):
            row = i + 1
            cov = round(d["passed"] / d["total"] * 100, 1) if d["total"] else 0
            ws.cell(row=row, column=1, value=mod)
            ws.cell(row=row, column=2, value=d["total"]).alignment = CENTER
            ws.cell(row=row, column=3, value=d["passed"]).alignment = CENTER
            ws.cell(row=row, column=4, value=d["failed"]).alignment = CENTER
            cov_cell = ws.cell(row=row, column=5, value=cov)
            cov_cell.alignment = CENTER
            cov_cell.number_format = "0.0"
            if cov >= 90:
                cov_cell.fill = PASS_FILL
            elif cov >= 70:
                cov_cell.fill = WARN_FILL
            else:
                cov_cell.fill = FAIL_FILL
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = BORDER

        # Bar chart
        if modules:
            chart = BarChart()
            chart.title = "Coverage by Module"
            chart.y_axis.title = "Coverage %"
            chart.x_axis.title = "Module"
            chart.width = 20
            data = Reference(ws, min_col=5, min_row=1, max_row=len(modules) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(modules) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"A{len(modules)+4}")

        for i, w in enumerate([20, 12, 10, 10, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet: Trends ──────────────────────────────────────────

    def _write_trends_sheet(self, ws, history: List[Dict]):
        ws.sheet_properties.tabColor = "9B59B6"
        headers = ["Run", "Date", "Total", "Passed", "Failed", "Pass Rate %", "Duration (s)"]
        self._write_header_row(ws, headers)

        for i, h in enumerate(history, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=h.get("run_id", f"Run {i}"))
            ws.cell(row=row, column=2, value=h.get("date", ""))
            ws.cell(row=row, column=3, value=h.get("total", 0)).alignment = CENTER
            ws.cell(row=row, column=4, value=h.get("passed", 0)).alignment = CENTER
            ws.cell(row=row, column=5, value=h.get("failed", 0)).alignment = CENTER
            ws.cell(row=row, column=6, value=h.get("pass_rate", 0)).alignment = CENTER
            ws.cell(row=row, column=7, value=h.get("duration_sec", 0)).alignment = CENTER
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = BORDER

    # ── Sheet: Gate ────────────────────────────────────────────

    def _write_gate_sheet(self, ws, gate: Dict):
        ws.sheet_properties.tabColor = "E67E22"
        ws.merge_cells("A1:D1")
        verdict = gate.get("verdict", "N/A")
        ws["A1"] = f"Release Gate: {verdict} (Score: {gate.get('score', 0)}%)"
        ws["A1"].font = Font(bold=True, size=14,
                             color="27AE60" if verdict == "PASS" else
                             "E74C3C" if verdict == "FAIL" else "F39C12")

        headers = ["Rule", "Actual", "Fail Threshold", "Warn Threshold", "Verdict", "Confidence"]
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=j, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

        for i, rule in enumerate(gate.get("rules", []), 1):
            row = i + 3
            ws.cell(row=row, column=1, value=rule.get("name", "")).border = BORDER
            ws.cell(row=row, column=2, value=rule.get("actual", 0)).border = BORDER
            ws.cell(row=row, column=3, value=rule.get("fail_threshold", 0)).border = BORDER
            ws.cell(row=row, column=4, value=rule.get("warn_threshold", 0)).border = BORDER
            v_cell = ws.cell(row=row, column=5, value=rule.get("verdict", ""))
            v_cell.border = BORDER
            v_cell.alignment = CENTER
            v_cell.font = Font(bold=True)
            v = rule.get("verdict", "")
            if v == "PASS":
                v_cell.fill = PASS_FILL
            elif v == "FAIL":
                v_cell.fill = FAIL_FILL
            else:
                v_cell.fill = WARN_FILL
            ws.cell(row=row, column=6, value=f"{rule.get('confidence', 0):.0%}").border = BORDER

        for i, w in enumerate([25, 12, 15, 15, 10, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── helpers ────────────────────────────────────────────────

    def _write_header_row(self, ws, headers: List[str]):
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER